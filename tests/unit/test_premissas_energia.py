"""Smoke tests for solar_bess_risk.premissas_energia (30y × month aggregation)."""

from __future__ import annotations

from dataclasses import replace

import numpy as np
import pytest

from solar_bess_risk.config import HOURS_PER_YEAR, SimulationParams
from solar_bess_risk.premissas_energia import (
    _MONTH_DAYS,
    _month_hour_bounds,
    aggregate_joint_injection_30y,
    export_premissas_energia,
    write_premissas_energia_xlsx,
)
from solar_bess_risk.profile import SolarProfile
from solar_bess_risk.simulation import ScenarioDefinition, simulate_scenario
from solar_bess_risk.data_sources import PriceProfile


def _daily_solar(peak_mw: float) -> np.ndarray:
    """Build an 8760 array with a simple bell-shaped daytime profile."""
    gen = np.zeros(HOURS_PER_YEAR)
    for h in range(HOURS_PER_YEAR):
        hour = h % 24
        if 6 <= hour <= 18:
            gen[h] = peak_mw * np.sin(np.pi * (hour - 6) / 12.0)
    return gen


@pytest.fixture
def params() -> SimulationParams:
    return SimulationParams(
        csv_path="/tmp/test.csv",
        mwac=100.0,
        capex_usd_per_kwh=200.0,
        usd_brl_rate=5.0,
        useful_life_years=30,
    )


@pytest.fixture
def solar_3yr() -> SolarProfile:
    base = _daily_solar(80.0)
    # 3 years with mild degradation so per-year selection is observable.
    years = np.stack([base * (1.0 - 0.01 * y) for y in range(3)])
    annual = float(base.sum())
    return SolarProfile(
        generation_mw=base,
        annual_energy_mwh=annual,
        fc=annual / (100.0 * HOURS_PER_YEAR),
        garantia_fisica_mw=30.0,
        csv_filename="t.csv",
        generation_lim_mw=base,
        generation_bess_mw=base,
        generation_years_lim_mw=years,
        generation_years_bess_mw=years,
        n_years=3,
    )


@pytest.fixture
def scenario() -> ScenarioDefinition:
    gf = 30.0
    return ScenarioDefinition(
        label="2025-4h",
        peak_hours=frozenset({18, 19, 20, 21}),
        duration_h=4,
        bess_power_mw=gf,
        bess_energy_mwh=gf * 4,
        capex_brl=gf * 4 * 200 * 1000 * 5.0,
    )


def test_month_bounds_sum_to_8760():
    bounds = _month_hour_bounds()
    assert len(bounds) == 13
    assert bounds[0] == 0
    assert bounds[-1] == HOURS_PER_YEAR
    assert sum(_MONTH_DAYS) == 365


def test_aggregate_shape_and_annual(solar_3yr, scenario, params):
    pld = np.full(HOURS_PER_YEAR, 200.0)
    result = aggregate_joint_injection_30y(
        solar=solar_3yr,
        pld=pld,
        price_source="pld_SE_2025",
        bq_submarket="SE",
        scenario=scenario,
        params=params,
        curtailment_series=None,
        rte_table={},
        start_year=2025,
        soh_table=None,
        must_mw=None,
        scenario_label="2025-4h",
    )
    assert result.monthly_mwh.shape == (30, 12)
    assert result.annual_mwh.shape == (30,)
    assert result.gf_annual_mw_med.shape == (30,)
    # Annual equals the row-wise monthly sum.
    np.testing.assert_allclose(
        result.annual_mwh, result.monthly_mwh.sum(axis=1), rtol=1e-9
    )
    # Annual GF is row-specific annual MWh / 8760.
    np.testing.assert_allclose(
        result.gf_annual_mw_med, result.annual_mwh / HOURS_PER_YEAR, rtol=1e-9
    )
    # GF P50 (MWmédio) equals mean annual / 8760.
    assert result.gf_p50_mw_med == pytest.approx(
        result.gf_annual_mw_med.mean()
    )
    # All energy is non-negative.
    assert (result.monthly_mwh >= -1e-9).all()


def test_aggregate_matches_direct_year1_dispatch(solar_3yr, scenario, params):
    """Year-1 monthly totals must match a direct simulate_scenario aggregation."""
    pld = np.full(HOURS_PER_YEAR, 200.0)
    result = aggregate_joint_injection_30y(
        solar=solar_3yr,
        pld=pld,
        price_source="pld_SE_2025",
        bq_submarket="SE",
        scenario=scenario,
        params=params,
        curtailment_series=None,
        rte_table={},
        start_year=2025,
        soh_table=None,
        must_mw=None,
    )
    price_profile = PriceProfile(pld, "pld_SE_2025", "SE", 2025)
    dispatch = simulate_scenario(
        solar_3yr, price_profile, scenario, params,
        curtailment_series=None, solar_year_idx=1, must_mw=None,
    )
    bounds = _month_hour_bounds()
    inj = np.asarray(dispatch.grid_injection_mwh, dtype=np.float64)
    expected = np.array([inj[bounds[m]:bounds[m + 1]].sum() for m in range(12)])
    np.testing.assert_allclose(result.monthly_mwh[0], expected, rtol=1e-9)


def test_monthly_export_is_below_solar_input_and_reconciles_rte_losses(
    solar_3yr, scenario, params
):
    """Monthly export must stay below source solar after BESS conversion losses."""
    pld = np.full(HOURS_PER_YEAR, 100.0)
    for day in range(365):
        start = day * 24
        pld[start + 10:start + 16] = 10.0
        pld[start + 18:start + 22] = 500.0

    lossy_scenario = replace(scenario, rte=0.80, charge_mode=3)
    result = aggregate_joint_injection_30y(
        solar=solar_3yr,
        pld=pld,
        price_source="synthetic_spread",
        bq_submarket="SE",
        scenario=lossy_scenario,
        params=params,
        curtailment_series=None,
        rte_table={},
        start_year=2025,
        soh_table=None,
        must_mw=None,
        n_years=2,
    )

    price_profile = PriceProfile(pld, "synthetic_spread", "SE", 2025)
    bounds = _month_hour_bounds()
    for offset in range(result.n_years):
        _, gen_bess = solar_3yr.get_year_arrays(offset + 1)
        dispatch = simulate_scenario(
            solar_3yr,
            price_profile,
            lossy_scenario,
            params,
            curtailment_series=None,
            solar_year_idx=offset + 1,
            must_mw=None,
        )

        for month in range(12):
            start, end = bounds[month], bounds[month + 1]
            solar_input = float(np.sum(gen_bess[start:end]))
            exported = float(result.monthly_mwh[offset, month])
            spilled = float(np.sum(dispatch.curtailment_lost_mwh[start:end]))
            charge = float(np.sum(dispatch.charge_mwh[start:end]))
            soc_before = float(dispatch.soc_mwh[start - 1]) if start else 0.0
            soc_after = float(dispatch.soc_mwh[end - 1])
            soc_delta = soc_after - soc_before
            rte_loss = charge * (1.0 - lossy_scenario.rte)

            assert exported < solar_input
            assert solar_input - exported == pytest.approx(
                spilled + rte_loss + soc_delta,
                abs=1e-6,
            )


def test_write_xlsx_layout(solar_3yr, scenario, params, tmp_path):
    import openpyxl

    pld = np.full(HOURS_PER_YEAR, 200.0)
    result = aggregate_joint_injection_30y(
        solar=solar_3yr,
        pld=pld,
        price_source="pld_SE_2025",
        bq_submarket="SE",
        scenario=scenario,
        params=params,
        curtailment_series=None,
        rte_table={},
        start_year=2025,
        scenario_label="2025-4h",
    )
    out = tmp_path / "premissas_energia.xlsx"
    path = write_premissas_energia_xlsx(result, out)
    wb = openpyxl.load_workbook(path, data_only=True)
    assert "Premissas-Energia" in wb.sheetnames
    ws = wb["Premissas-Energia"]
    # Header row (row 2): "Ano / Mês" | 1..12 | Anual | GF
    assert ws.cell(row=2, column=1).value == "Ano / Mês"
    assert [ws.cell(row=2, column=2 + m).value for m in range(12)] == list(range(1, 13))
    assert ws.cell(row=2, column=14).value == "Anual"
    assert ws.cell(row=2, column=15).value == "GF"
    # 30 data rows starting at row 3, year labels 1..30
    assert ws.cell(row=3, column=1).value == 1
    assert ws.cell(row=32, column=1).value == 30
    # Annual cell equals sum of the 12 monthly cells on that row.
    monthly = [ws.cell(row=3, column=2 + m).value for m in range(12)]
    assert ws.cell(row=3, column=14).value == pytest.approx(sum(monthly))
    # GF mirrors the financial model block: ROUND(annual MWh / 8760, 2) per row.
    assert ws.cell(row=3, column=15).value == pytest.approx(
        round(ws.cell(row=3, column=14).value / HOURS_PER_YEAR, 2)
    )
    assert ws.cell(row=4, column=15).value == pytest.approx(
        round(ws.cell(row=4, column=14).value / HOURS_PER_YEAR, 2)
    )
    assert ws.cell(row=3, column=15).value != pytest.approx(result.gf_p50_mw_med)


def test_aggregate_uses_ano1_curve_for_first_operational_year(
    solar_3yr, scenario, params, monkeypatch
):
    """The first exported year skips supplier Ano 0 and starts at Ano 1."""
    import types

    import solar_bess_risk.premissas_energia as pe

    calls: list[tuple[float, float, int]] = []

    def fake_simulate_scenario(
        solar,
        prices,
        yearly_scenario,
        params,
        *,
        curtailment_series=None,
        solar_year_idx=1,
        must_mw=None,
    ):
        calls.append(
            (
                yearly_scenario.rte,
                yearly_scenario.bess_energy_mwh,
                solar_year_idx,
            )
        )
        return types.SimpleNamespace(grid_injection_mwh=np.ones(HOURS_PER_YEAR))

    monkeypatch.setattr(pe, "simulate_scenario", fake_simulate_scenario)

    aggregate_joint_injection_30y(
        solar=solar_3yr,
        pld=np.full(HOURS_PER_YEAR, 200.0),
        price_source="pld_SE_2025",
        bq_submarket="SE",
        scenario=scenario,
        params=params,
        curtailment_series=None,
        rte_table={2025: 0.90, 2026: 0.80, 2027: 0.70},
        soh_table={2025: 1.0, 2026: 0.95, 2027: 0.90},
        start_year=2025,
        n_years=2,
    )

    assert calls == pytest.approx(
        [
            (0.80, scenario.bess_energy_mwh * 0.95, 1),
            (0.70, scenario.bess_energy_mwh * 0.90, 2),
        ]
    )

    calls.clear()
    aggregate_joint_injection_30y(
        solar=solar_3yr,
        pld=np.full(HOURS_PER_YEAR, 200.0),
        price_source="pld_SE_2025",
        bq_submarket="SE",
        scenario=scenario,
        params=params,
        curtailment_series=None,
        rte_table={},
        rte_fallback=0.77,
        start_year=2025,
        n_years=1,
    )

    assert calls == pytest.approx([(0.77, scenario.bess_energy_mwh, 1)])


def test_export_returns_none_when_requested_scenario_missing(
    solar_3yr, scenario, params, tmp_path
):
    """Avoid silently exporting the wrong tab when 2025-4h is unavailable."""
    out = tmp_path / "premissas_energia.xlsx"

    result = export_premissas_energia(
        results_by_key={
            "2025-2h": (
                None,
                np.full(HOURS_PER_YEAR, 200.0),
                solar_3yr.garantia_fisica_mw,
                solar_3yr.generation_lim_mw,
                scenario.peak_hours,
                2,
                2025,
                scenario.rte,
                scenario,
            )
        },
        solar=solar_3yr,
        params=params,
        pld_by_year={2025: np.full(HOURS_PER_YEAR, 200.0)},
        price_sources_by_year={2025: "pld_SE_2025"},
        curtailment_enabled=False,
        rte_table={2026: 0.80},
        rte_fallback=0.80,
        output_path=out,
        scenario_key="2025-4h",
    )

    assert result is None
    assert not out.exists()
