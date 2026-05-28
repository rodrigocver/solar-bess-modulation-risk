"""Integration tests: end-to-end run with mocked BigQuery."""

from __future__ import annotations

import json
import os
from pathlib import Path
from unittest.mock import patch, MagicMock

import numpy as np
import pytest

from solar_bess_risk.config import HOURS_PER_YEAR, SCENARIO_TEMPLATES, SimulationParams
from solar_bess_risk.simulation import ScenarioDefinition


def _build_scenarios(gf: float, params: SimulationParams) -> list[ScenarioDefinition]:
    """Build ScenarioDefinition list from templates."""
    capex_per_mwh = params.capex_usd_per_kwh * 1000 * params.usd_brl_rate
    return [
        ScenarioDefinition(
            label=t.label,
            peak_hours=t.peak_hours,
            duration_h=t.duration_h,
            bess_power_mw=gf,
            bess_energy_mwh=gf * t.duration_h,
            capex_brl=gf * t.duration_h * capex_per_mwh,
            charge_power_mw=gf,
        )
        for t in SCENARIO_TEMPLATES
    ]


@pytest.fixture
def tmp_csv(tmp_path):
    """Create a minimal solar CSV for testing."""
    gen = np.clip(np.sin(np.linspace(0, 2 * np.pi * 365, HOURS_PER_YEAR)) * 300 + 100, 0, None)
    csv_path = tmp_path / "test_solar.csv"
    csv_path.write_text("\n".join(f"{v:.2f}" for v in gen))
    return str(csv_path)


@pytest.fixture
def mock_bq_prices():
    """Return uniform prices for mocking BigQuery."""
    return np.full(HOURS_PER_YEAR, 100.0)


@pytest.fixture
def params(tmp_csv):
    """Create test SimulationParams."""
    return SimulationParams(
        csv_path=tmp_csv,
        mwac=600.0,
        bq_year=2025,
        bq_submarket="SE",
        capex_usd_per_kwh=200.0,
        usd_brl_rate=5.0,
        useful_life_years=20,
        bq_service_account_path=None,
    )


class TestFullRun:
    """End-to-end integration tests."""

    def test_produces_3_scenario_results(self, params, mock_bq_prices):
        """Full pipeline produces exactly 3 ScenarioResult objects."""
        from solar_bess_risk.profile import load_solar_csv
        from solar_bess_risk.data_sources import PriceProfile
        from solar_bess_risk.simulation import simulate_all_scenarios
        from solar_bess_risk.economics import compute_all_scenarios

        solar = load_solar_csv(params.csv_path, params.mwac)
        prices = PriceProfile(
            prices_brl_per_mwh=mock_bq_prices,
            source="test_mock",
            bq_submarket="SE",
            bq_year=2025,
        )
        scenarios = _build_scenarios(solar.garantia_fisica_mw, params)
        dispatch_pairs = simulate_all_scenarios(solar, prices, scenarios, params)
        results = compute_all_scenarios(solar, prices, dispatch_pairs, params)

        assert len(results) == 2
        labels = {r.scenario.label for r in results}
        assert labels == {"A", "B"}

    def test_produces_html_report(self, params, mock_bq_prices, tmp_path):
        """Full pipeline produces report.html."""
        from solar_bess_risk.profile import load_solar_csv
        from solar_bess_risk.data_sources import PriceProfile
        from solar_bess_risk.simulation import simulate_all_scenarios
        from solar_bess_risk.economics import compute_all_scenarios
        from solar_bess_risk.report_export import write_report

        solar = load_solar_csv(params.csv_path, params.mwac)
        prices = PriceProfile(
            prices_brl_per_mwh=mock_bq_prices,
            source="test_mock",
            bq_submarket="SE",
            bq_year=2025,
        )
        scenarios = _build_scenarios(solar.garantia_fisica_mw, params)
        dispatch_pairs = simulate_all_scenarios(solar, prices, scenarios, params)
        results = compute_all_scenarios(solar, prices, dispatch_pairs, params)

        output_dir = tmp_path / "output"
        report_path = write_report(results, prices, params, solar, output_dir)
        assert Path(report_path).exists()
        content = Path(report_path).read_text()
        assert "Cenário" in content
        assert "Portaria MME 101/2016" in content

    def test_builds_backtest_html_report_with_submarket_and_rte_metadata(self, tmp_path):
        """Backtest output writer creates the main report.html artifact."""
        from solar_bess_risk.report_excel import build_html_report
        from solar_bess_risk.simulation import DispatchResult

        dispatch = DispatchResult(
            soc_mwh=np.zeros(HOURS_PER_YEAR),
            charge_mwh=np.zeros(HOURS_PER_YEAR),
            discharge_mwh=np.zeros(HOURS_PER_YEAR),
            grid_injection_mwh=np.zeros(HOURS_PER_YEAR),
            deficit_mwh=np.ones(HOURS_PER_YEAR),
            residual_deficit_mwh=np.ones(HOURS_PER_YEAR),
            curtailment_mwh=np.zeros(HOURS_PER_YEAR),
            curtailment_lost_mwh=np.zeros(HOURS_PER_YEAR),
        carga_nao_realizada_diaria_mwh=np.zeros(365),
        )
        results_by_key = {
            "2025-2h": (
                dispatch,
                np.full(HOURS_PER_YEAR, 100.0),
                25.0,
                np.zeros(HOURS_PER_YEAR),
                frozenset({18, 19}),
                2,
                2025,
                0.8625,
            )
        }

        report_path = build_html_report(
            results_by_key,
            tmp_path / "report.html",
            mwac=100.0,
            usd_brl_rate=5.0,
            bq_submarket="NE",
            rte_metadata={
                "rte_source_file": "11 - Envision.xlsx",
                "typical_block_mwh": 10.1,
                "pcs_mva": 2.52,
            },
        )
        content = Path(report_path).read_text(encoding="utf-8")
        assert Path(report_path).name == "report.html"
        assert "Submercado PLD:</strong> NE" in content
        assert "typical_block_mwh" in content
        assert "10.1" in content
        assert "pcs_mva" in content
        assert "2.52" in content

    def test_builds_consultancy_report_with_net_balance_economics(self, tmp_path):
        """Consultancy report must build KPI/payback after net-balance economy is computed."""
        from solar_bess_risk.report_consultancy import build_consultancy_report
        from solar_bess_risk.simulation import DispatchResult, ScenarioDefinition

        generation = np.full(HOURS_PER_YEAR, 50.0)
        charge = np.zeros(HOURS_PER_YEAR)
        discharge = np.zeros(HOURS_PER_YEAR)
        grid = generation.copy()
        deficit = np.zeros(HOURS_PER_YEAR)
        residual = np.zeros(HOURS_PER_YEAR)
        pld = np.full(HOURS_PER_YEAR, 100.0)

        charge[10::24] = 10.0
        discharge[20::24] = 10.0
        grid = generation - charge + discharge
        pld[20::24] = 1000.0
        scenario = ScenarioDefinition(
            label="A",
            peak_hours=frozenset({18, 19}),
            duration_h=2,
            bess_power_mw=10.0,
            bess_energy_mwh=20.0,
            capex_brl=1_000_000.0,
            charge_power_mw=10.0,
        )
        dispatch = DispatchResult(
            soc_mwh=np.zeros(HOURS_PER_YEAR),
            charge_mwh=charge,
            discharge_mwh=discharge,
            grid_injection_mwh=grid,
            deficit_mwh=deficit,
            residual_deficit_mwh=residual,
            curtailment_mwh=np.zeros(HOURS_PER_YEAR),
            curtailment_lost_mwh=np.zeros(HOURS_PER_YEAR),
            carga_nao_realizada_diaria_mwh=np.zeros(365),
        )
        results_by_key = {
            "2025-2h": (
                dispatch,
                pld,
                50.0,
                generation,
                scenario.peak_hours,
                scenario.duration_h,
                2025,
                1.0,
                scenario,
            )
        }

        path = build_consultancy_report(
            results_by_key,
            tmp_path / "relatorio_diretoria.html",
            mwac=100.0,
            usd_brl_rate=5.0,
            bq_submarket="SE",
            garantia_fisica_mw=50.0,
            fc=0.5,
            charge_mode=3,
        )

        content = Path(path).read_text(encoding="utf-8")
        assert "Relatório Executivo" in content
        assert "Economia Anual" in content

    def test_run_manifest_tracks_selected_submarket_and_executed_scenarios(self, params):
        """Main manifest helper preserves non-SE submarket labels and full scenario sizing."""
        from dataclasses import replace

        from solar_bess_risk.__main__ import _build_run_manifest
        from solar_bess_risk.profile import SolarProfile
        from solar_bess_risk.simulation import DispatchResult

        ne_params = replace(params, bq_submarket="NE")
        solar = SolarProfile(
            generation_mw=np.zeros(HOURS_PER_YEAR),
            annual_energy_mwh=100.0,
            fc=0.01,
            garantia_fisica_mw=25.0,
            csv_filename="solar.csv",
        )
        dispatch = DispatchResult(
            soc_mwh=np.zeros(HOURS_PER_YEAR),
            charge_mwh=np.zeros(HOURS_PER_YEAR),
            discharge_mwh=np.zeros(HOURS_PER_YEAR),
            grid_injection_mwh=np.zeros(HOURS_PER_YEAR),
            deficit_mwh=np.ones(HOURS_PER_YEAR),
            residual_deficit_mwh=np.ones(HOURS_PER_YEAR),
            curtailment_mwh=np.zeros(HOURS_PER_YEAR),
            curtailment_lost_mwh=np.zeros(HOURS_PER_YEAR),
        carga_nao_realizada_diaria_mwh=np.zeros(365),
        )
        results_by_key = {
            "2025-2h": (
                dispatch,
                np.full(HOURS_PER_YEAR, 100.0),
                25.0,
                np.zeros(HOURS_PER_YEAR),
                frozenset({18, 19}),
                2,
                2025,
                0.8625,
            )
        }

        manifest = _build_run_manifest(
            run_id="20260520-120000-abcdef0",
            params=ne_params,
            solar=solar,
            results_by_key=results_by_key,
            price_sources_by_year={2025: "bigquery_pld_NE_2025"},
            rte_path="dados/11 - Envision.xlsx",
            rte_table={2025: 0.8625},
            rte_acum=0.8625,
            curtailment_enabled=False,
            rte_metadata={"typical_block_mwh": 10.1, "pcs_mva": 2.52},
        )

        assert manifest.price_source == "bigquery_pld_NE_multi_year"
        assert manifest.params["bq_submarket"] == "NE"
        assert manifest.price_sources_by_year == {"2025": "bigquery_pld_NE_2025"}
        assert manifest.scenarios[0]["bess_power_mw"] == 25.0
        assert manifest.scenarios[0]["bess_energy_mwh"] == 50.0
        assert manifest.rte["metadata"]["pcs_mva"] == 2.52

    def test_produces_manifest_json(self, params, mock_bq_prices, tmp_path):
        """Full pipeline produces manifest.json."""
        from solar_bess_risk.profile import load_solar_csv
        from solar_bess_risk.data_sources import PriceProfile
        from solar_bess_risk.simulation import simulate_all_scenarios
        from solar_bess_risk.economics import compute_all_scenarios
        from solar_bess_risk.manifest import RunManifest, generate_run_id, hash_params, write_manifest

        solar = load_solar_csv(params.csv_path, params.mwac)
        prices = PriceProfile(
            prices_brl_per_mwh=mock_bq_prices,
            source="test_mock",
            bq_submarket="SE",
            bq_year=2025,
        )
        scenarios = _build_scenarios(solar.garantia_fisica_mw, params)
        dispatch_pairs = simulate_all_scenarios(solar, prices, scenarios, params)
        results = compute_all_scenarios(solar, prices, dispatch_pairs, params)

        output_dir = tmp_path / "output"
        output_dir.mkdir(parents=True, exist_ok=True)

        manifest = RunManifest(
            tool_version="2.0.0",
            run_id=generate_run_id(),
            timestamp_iso8601="2025-01-01T00:00:00Z",
            params_sha256=hash_params(params),
            profile_source=solar.csv_filename,
            price_source=prices.source,
            fc=solar.fc,
            garantia_fisica_mw=solar.garantia_fisica_mw,
            scenarios=[r.scenario.label for r in results],
        )
        manifest_path = write_manifest(manifest, output_dir)
        data = json.loads(Path(manifest_path).read_text())
        assert data["scenarios"] == ["A", "B"]
        assert data["tool_version"] == "2.0.0"

    def test_scenario_labels_are_a_b_for_2h_and_4h_products(self, params, mock_bq_prices):
        """Final product scenarios use labels A and B for 2h and 4h products."""
        from solar_bess_risk.profile import load_solar_csv
        from solar_bess_risk.data_sources import PriceProfile
        from solar_bess_risk.simulation import simulate_all_scenarios
        from solar_bess_risk.economics import compute_all_scenarios

        solar = load_solar_csv(params.csv_path, params.mwac)
        prices = PriceProfile(
            prices_brl_per_mwh=mock_bq_prices,
            source="test_mock",
            bq_submarket="SE",
            bq_year=2025,
        )
        scenarios = _build_scenarios(solar.garantia_fisica_mw, params)
        dispatch_pairs = simulate_all_scenarios(solar, prices, scenarios, params)
        results = compute_all_scenarios(solar, prices, dispatch_pairs, params)

        # SC-001 = A (2h), SC-002 = B (4h)
        durations = {r.scenario.label: r.scenario.duration_h for r in results}
        assert durations == {"A": 2, "B": 4}

    def test_excel_report_places_scenario_metadata_once_and_daily_missed_charge(self, tmp_path):
        """Excel report places scenario metadata on first data row and missed charge daily."""
        from openpyxl import load_workbook

        from solar_bess_risk.projection import CashflowProjection
        from solar_bess_risk.report_excel import build_excel_report
        from solar_bess_risk.simulation import DispatchResult, ScenarioDefinition

        dispatch = DispatchResult(
            soc_mwh=np.zeros(HOURS_PER_YEAR),
            charge_mwh=np.zeros(HOURS_PER_YEAR),
            discharge_mwh=np.zeros(HOURS_PER_YEAR),
            grid_injection_mwh=np.zeros(HOURS_PER_YEAR),
            deficit_mwh=np.ones(HOURS_PER_YEAR),
            residual_deficit_mwh=np.ones(HOURS_PER_YEAR),
            curtailment_mwh=np.zeros(HOURS_PER_YEAR),
            curtailment_lost_mwh=np.zeros(HOURS_PER_YEAR),
            carga_nao_realizada_diaria_mwh=np.arange(365, dtype=float),
        )
        scenario = ScenarioDefinition(
            label="A",
            peak_hours=frozenset({18, 19}),
            duration_h=2,
            bess_power_mw=12.0,
            bess_energy_mwh=99.0,
            capex_brl=123_456_789.0,
            charge_power_mw=12.0,
        )
        projection = CashflowProjection(
            payback_years=7.25,
            lifetime_net_savings_brl=10_000_000.0,
            lcos_brl_per_mwh=321.45,
            lifetime_discharge_mwh=123_000.0,
            annual_gross_savings_brl=(1_000_000.0,),
            annual_net_savings_brl=(900_000.0,),
            annual_discharge_mwh=(6_150.0,),
            annual_rte=(0.85,),
        )
        results_by_key = {
            "2025-2h": (
                dispatch,
                np.full(HOURS_PER_YEAR, 100.0),
                10.0,
                np.zeros(HOURS_PER_YEAR),
                scenario.peak_hours,
                scenario.duration_h,
                2025,
                0.85,
                scenario,
                projection,
            )
        }

        import pandas as pd

        block_detail = pd.DataFrame([
            {
                "cenario": "2025-2h",
                "n_blocos": 1,
                "roi_vida_util": 1.2,
                "ranking_retorno": 1,
                "ranking_payback": 1,
                "recomendado": True,
            }
        ])
        block_recommended = block_detail.copy()
        path = build_excel_report(
            results_by_key,
            tmp_path / "report.xlsx",
            100.0,
            5.0,
            block_optimization=(block_detail, block_recommended),
        )
        wb = load_workbook(path, data_only=True)
        assert "diagnostico_carga" in wb.sheetnames
        assert "diagnostico_diario" in wb.sheetnames
        assert "otimizacao_blocos" in wb.sheetnames
        assert "otimizacao_recomendada" in wb.sheetnames
        ws = wb["2025-2h"]
        headers = [cell.value for cell in ws[1]]
        first_data_row = 2
        second_data_row = 3
        capex_col = headers.index("capex_brl") + 1
        power_col = headers.index("bess_power_mw") + 1
        energy_col = headers.index("bess_energy_mwh") + 1
        mode_col = headers.index("modo_operacao") + 1
        rte_col = headers.index("rte") + 1
        missed_col = headers.index("carga_nao_realizada_mwh_dia") + 1
        payback_col = headers.index("payback_anos") + 1
        lcos_col = headers.index("lcos_brl_mwh") + 1
        lifetime_discharge_col = headers.index("descarga_bess_mwh_vida_util") + 1
        injection_sem_col = headers.index("injecao_sem_bess_mwh") + 1
        injection_com_col = headers.index("injecao_com_bess_mwh") + 1

        assert ws.cell(first_data_row, capex_col).value == round(scenario.capex_brl)
        assert ws.cell(first_data_row, power_col).value == round(scenario.bess_power_mw)
        assert ws.cell(first_data_row, energy_col).value == round(scenario.bess_energy_mwh)
        assert ws.cell(first_data_row, mode_col).value is not None
        assert ws.cell(first_data_row, rte_col).value == 0.85

        assert ws.cell(second_data_row, capex_col).value is None
        assert ws.cell(second_data_row, power_col).value is None
        assert ws.cell(second_data_row, energy_col).value is None
        assert ws.cell(second_data_row, mode_col).value is None
        assert ws.cell(second_data_row, rte_col).value is None

        assert ws.cell(25, missed_col).value == 0
        assert ws.cell(49, missed_col).value == 1
        assert ws.cell(8762, payback_col).value == 7.25
        assert ws.cell(8762, lcos_col).value == 321.45
        assert ws.cell(8762, lifetime_discharge_col).value == 123000
        assert ws.cell(first_data_row, injection_sem_col).value == 0
        assert ws.cell(first_data_row, injection_com_col).value == 0

        diag = wb["diagnostico_carga"]
        diag_headers = [cell.value for cell in diag[1]]
        assert "curtailment_pct_geracao" in diag_headers
        assert "sem_spread_economico_mwh" in diag_headers
        assert "carga_nao_realizada_pct_ciclo_teorico" in diag_headers
        assert "gargalo_potencia_carga_mwh" in diag_headers
        assert "gargalo_janela_descarga_mwh" in diag_headers
        assert "potencial_extensao_d1_total_mwh" in diag_headers

    def test_builds_block_optimization_html_report(self, tmp_path):
        """Block optimization output creates a dedicated HTML report."""
        import pandas as pd

        from solar_bess_risk.report_optimization import build_block_optimization_html

        detail = pd.DataFrame([
            {
                "cenario": "2026-4h",
                "capex_scenario": "capex_-10%",
                "n_blocos": 80,
                "multiplo_blocos_gf": 1.1,
                "bess_power_mw": 201.6,
                "bess_energy_mwh": 808.0,
                "capex_brl": 1_000_000_000.0,
                "economia_liquida_anual_brl": 150_000_000.0,
                "payback_anos": 6.7,
                "lcos_brl_mwh": 300.0,
                "roi_vida_util": 1.5,
                "ranking_retorno": 1,
                "ranking_payback": 1,
                "projecao_rte_completa": True,
                "recomendado": True,
            }
        ])
        recommended = detail.copy()

        path = build_block_optimization_html(
            detail,
            recommended,
            tmp_path / "otimizacao_blocos.html",
        )

        content = Path(path).read_text(encoding="utf-8")
        assert "Otimização de Blocos BESS" in content
        assert "capex_-10%" in content
        assert "Recomendação por Cenário e CAPEX" in content
