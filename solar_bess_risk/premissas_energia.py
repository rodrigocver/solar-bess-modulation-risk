"""Premissas-Energia export — joint solar+BESS dispatch aggregated 30y × month.

This module produces a standalone ``.xlsx`` with a ``Premissas-Energia`` sheet
mirroring the layout used by the financial-modelling team (block
"Geração P50 PMI Complexo (MWh)"): a 30-year × 12-month matrix of the executed
grid injection under the joint solar + BESS dispatch, plus an annual total and
the garantia-física (GF) column.

The solar series already carries 30 years of degradation
(``SolarProfile.generation_years_bess_mw``); this module only applies the BESS
SOH/RTE per calendar year, re-simulating each year exactly like
``projection.project_cashflows_with_rte`` and aggregating
``grid_injection_mwh`` month by month.

Public API
----------
aggregate_joint_injection_30y(...) -> PremissasEnergiaResult
write_premissas_energia_xlsx(result, output_path, ...) -> str
export_premissas_energia(...) -> str | None
"""

from __future__ import annotations

from dataclasses import dataclass, replace
from pathlib import Path

import numpy as np

from solar_bess_risk.config import (
    DEFAULT_RTE_COMMISSIONING_YEAR,
    HOURS_PER_YEAR,
    SimulationParams,
)
from solar_bess_risk.data_sources import PriceProfile
from solar_bess_risk.profile import SolarProfile
from solar_bess_risk.projection import (
    MAX_BESS_CALENDAR_YEARS,
    rte_for_year,
    soh_for_year,
)
from solar_bess_risk.simulation import ScenarioDefinition, simulate_scenario

# Month boundaries for a non-leap 365-day year (8760 hours).
_MONTH_DAYS = (31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31)
_MONTH_LABELS = (
    "jan", "fev", "mar", "abr", "mai", "jun",
    "jul", "ago", "set", "out", "nov", "dez",
)


def _month_hour_bounds() -> list[int]:
    """Return the 13 cumulative hour offsets delimiting the 12 calendar months."""
    bounds = [0]
    for days in _MONTH_DAYS:
        bounds.append(bounds[-1] + days * 24)
    if bounds[-1] != HOURS_PER_YEAR:  # defensive: keep aligned with 8760
        bounds[-1] = HOURS_PER_YEAR
    return bounds


@dataclass(frozen=True)
class PremissasEnergiaResult:
    """Aggregated joint solar+BESS dispatch over the project lifetime."""

    monthly_mwh: np.ndarray            # shape (n_years, 12)
    annual_mwh: np.ndarray             # shape (n_years,)
    gf_annual_mw_med: np.ndarray       # shape (n_years,), annual GF in MWmédio
    gf_p50_mw_med: float               # average annual GF over the exported years
    n_years: int
    scenario_label: str


def aggregate_joint_injection_30y(
    *,
    solar: SolarProfile,
    pld: np.ndarray,
    price_source: str,
    bq_submarket: str,
    scenario: ScenarioDefinition,
    params: SimulationParams,
    curtailment_series: np.ndarray | None,
    rte_table: dict[int, float],
    start_year: int,
    rte_fallback: float | None = None,
    soh_table: dict[int, float] | None = None,
    must_mw: float | None = None,
    scenario_label: str = "",
    n_years: int = MAX_BESS_CALENDAR_YEARS,
) -> PremissasEnergiaResult:
    """Aggregate the joint solar+BESS grid injection into a 30y × 12-month matrix.

    Re-simulates each operational year with that year's solar degradation
    (via ``solar_year_idx``) and BESS SOH/RTE, mirroring
    ``project_cashflows_with_rte``. The executed grid injection
    (``dispatch.grid_injection_mwh``) is then summed per calendar month.

    Returns
    -------
    PremissasEnergiaResult
        Monthly/annual MWh matrices and the P50 GF (MWmédio).
    """
    if rte_fallback is not None:
        fallback_rte = rte_fallback
    else:
        fallback_rte = (
            scenario.rte if scenario.rte != 1.0 else params.bess_roundtrip_efficiency
        )
    if start_year < DEFAULT_RTE_COMMISSIONING_YEAR:
        start_year = DEFAULT_RTE_COMMISSIONING_YEAR

    price_profile = PriceProfile(pld, price_source, bq_submarket, start_year)
    bounds = _month_hour_bounds()

    monthly = np.zeros((n_years, 12), dtype=np.float64)

    for offset in range(n_years):
        # Operational Year 1 corresponds to Ano 1 of the supplier curve, matching
        # projection.project_cashflows_with_rte.
        cashflow_year = start_year + offset + 1
        rte = rte_for_year(rte_table, cashflow_year, fallback_rte)
        soh = soh_for_year(soh_table, cashflow_year, 1.0)
        yearly_energy_mwh = scenario.bess_energy_mwh * soh
        yearly_scenario = replace(scenario, rte=rte, bess_energy_mwh=yearly_energy_mwh)
        solar_year_idx = min(offset + 1, solar.n_years)

        dispatch = simulate_scenario(
            solar,
            price_profile,
            yearly_scenario,
            params,
            curtailment_series=curtailment_series,
            solar_year_idx=solar_year_idx,
            must_mw=must_mw,
        )

        injection = np.asarray(dispatch.grid_injection_mwh, dtype=np.float64)
        for m in range(12):
            monthly[offset, m] = injection[bounds[m]:bounds[m + 1]].sum()

    annual = monthly.sum(axis=1)
    gf_annual_mw_med = (
        annual / HOURS_PER_YEAR if n_years else np.array([], dtype=np.float64)
    )
    gf_p50_mw_med = float(gf_annual_mw_med.mean()) if n_years else 0.0

    return PremissasEnergiaResult(
        monthly_mwh=monthly,
        annual_mwh=annual,
        gf_annual_mw_med=gf_annual_mw_med,
        gf_p50_mw_med=gf_p50_mw_med,
        n_years=n_years,
        scenario_label=scenario_label,
    )


def write_premissas_energia_xlsx(
    result: PremissasEnergiaResult,
    output_path: str | Path,
    *,
    sheet_name: str = "Premissas-Energia",
    title: str | None = None,
) -> str:
    """Write the aggregated matrix to an ``.xlsx`` mirroring the financial layout.

    Layout (block "Geração P50 PMI Complexo (MWh)"):
        title row
        header: "Ano / Mês" | 1 | 2 | ... | 12 | Anual | GF
        rows:   year(1..N)  | monthly MWh ...   | total | annual GF(MWmédio)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if title is None:
        suffix = f" — cenário {result.scenario_label}" if result.scenario_label else ""
        title = f"Geração Conjunta Solar + BESS (MWh){suffix}"

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    # Title row
    ws.cell(row=1, column=1, value=title).font = bold

    # Header row
    header_row = 2
    ws.cell(row=header_row, column=1, value="Ano / Mês").font = bold
    for m in range(12):
        c = ws.cell(row=header_row, column=2 + m, value=m + 1)
        c.font = bold
        c.alignment = center
    annual_col = 2 + 12          # column 14
    gf_col = annual_col + 1      # column 15
    ws.cell(row=header_row, column=annual_col, value="Anual").font = bold
    ws.cell(row=header_row, column=gf_col, value="GF").font = bold

    # Data rows
    for year_idx in range(result.n_years):
        r = header_row + 1 + year_idx
        ws.cell(row=r, column=1, value=year_idx + 1)
        for m in range(12):
            ws.cell(row=r, column=2 + m, value=float(result.monthly_mwh[year_idx, m]))
        ws.cell(row=r, column=annual_col, value=float(result.annual_mwh[year_idx]))
        # Match the financial model block: row GF = ROUND(annual MWh / 8760, 2).
        ws.cell(
            row=r,
            column=gf_col,
            value=round(float(result.gf_annual_mw_med[year_idx]), 2),
        )

    # Widen the label column for readability.
    ws.column_dimensions["A"].width = 12

    wb.save(str(output_path))
    return str(output_path)


def export_premissas_energia(
    *,
    results_by_key: dict[str, tuple],
    solar: SolarProfile,
    params: SimulationParams,
    pld_by_year: dict[int, np.ndarray],
    price_sources_by_year: dict[int, str],
    curtailment_enabled: bool,
    rte_table: dict[int, float],
    rte_fallback: float,
    output_path: str | Path,
    soh_table: dict[int, float] | None = None,
    scenario_key: str = "2025-4h",
) -> str | None:
    """Build the Premissas-Energia workbook for a chosen scenario tab.

    Picks ``scenario_key`` (default ``"2025-4h"``) from ``results_by_key`` — the
    "sem redução de MUST" tab — and exports the joint solar+BESS dispatch
    aggregated 30y × month. Returns the output path, or ``None`` when the
    scenario tab is unavailable.
    """
    from solar_bess_risk.curtailment import get_curtailment_for_scenario

    data = results_by_key.get(scenario_key)
    if data is None:
        return None

    # results_by_key tuple layout (see __main__): index 1=pld, 6=year, 8=scenario.
    scenario = data[8] if len(data) > 8 else None
    year = data[6]
    if scenario is None:
        return None

    pld = pld_by_year[year]
    price_source = price_sources_by_year.get(
        year, f"pld_{params.bq_submarket}_{year}"
    )
    gen_lim = (
        solar.generation_lim_mw
        if solar.generation_lim_mw is not None
        else solar.generation_mw
    )
    curtailment_series = get_curtailment_for_scenario(
        year,
        curtailment_enabled,
        gen_lim,
        path=params.curtailment_path,
        factor_2026=params.curtailment_factor_2026,
        factor_2025=params.curtailment_factor_2025,
    )

    result = aggregate_joint_injection_30y(
        solar=solar,
        pld=pld,
        price_source=price_source,
        bq_submarket=params.bq_submarket,
        scenario=scenario,
        params=params,
        curtailment_series=curtailment_series,
        rte_table=rte_table,
        start_year=year,
        rte_fallback=rte_fallback,
        soh_table=soh_table,
        must_mw=None,  # the "2025-4h" tab is simulated sem redução de MUST
        scenario_label=scenario_key,
    )

    return write_premissas_energia_xlsx(result, output_path)
